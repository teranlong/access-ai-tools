"""
xlsx_to_md.py — Convert .xlsx/.xls to CSV per sheet + Markdown summary.

Uses openpyxl for .xlsx and xlrd for .xls (Excel 97-2003 binary).
"""

from __future__ import annotations

import csv
import io
from pathlib import Path

from .base import ConversionResult, make_output_path, make_output_dir


def _safe_cell(value) -> str:
    """Convert any cell value to a safe CSV string."""
    if value is None:
        return ""
    return str(value)


def _xlsx_to_sheets(source: Path) -> list[tuple[str, list[list[str]]]]:
    """Return list of (sheet_name, rows) from an .xlsx file."""
    import openpyxl
    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    result = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [[_safe_cell(cell.value) for cell in row] for row in ws.iter_rows()]
        result.append((name, rows))
    wb.close()
    return result


def _xls_to_sheets(source: Path) -> list[tuple[str, list[list[str]]]]:
    """Return list of (sheet_name, rows) from a legacy .xls file."""
    import xlrd
    wb = xlrd.open_workbook(str(source))
    result = []
    for sheet in wb.sheets():
        rows = []
        for r in range(sheet.nrows):
            row = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    row.append("")
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    # Format dates readably
                    dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                    row.append(dt.strftime("%Y-%m-%d"))
                else:
                    row.append(str(cell.value))
            rows.append(row)
        result.append((sheet.name, rows))
    return result


def _make_md_preview(rows: list[list[str]], max_rows: int = 5) -> str:
    """Generate a Markdown table preview from the first few rows."""
    if not rows:
        return "_Empty sheet_"
    header = rows[0]
    sep = ["---"] * len(header)
    body = rows[1 : max_rows + 1]

    def fmt_row(r: list[str]) -> str:
        cells = " | ".join(c.replace("|", "\\|").replace("\n", " ") for c in r)
        return f"| {cells} |"

    lines = [fmt_row(header), fmt_row(sep)]
    lines += [fmt_row(r) for r in body]
    if len(rows) - 1 > max_rows:
        lines.append(f"\n_... {len(rows) - 1 - max_rows} more rows_")
    return "\n".join(lines)


def convert(
    source: Path,
    output_root: Path,
    project_root: Path,
    **kwargs,
) -> ConversionResult:
    """Convert spreadsheet to per-sheet CSVs and a Markdown summary."""
    ext = source.suffix.lower()

    try:
        if ext == ".xls":
            try:
                sheets = _xls_to_sheets(source)
            except ImportError as exc:
                return ConversionResult(
                    success=False,
                    action="error",
                    error=f"Missing dependency: {exc}. Run: pip install xlrd",
                )
        else:
            try:
                sheets = _xlsx_to_sheets(source)
            except ImportError as exc:
                return ConversionResult(
                    success=False,
                    action="error",
                    error=f"Missing dependency: {exc}. Run: pip install openpyxl",
                )
            except Exception as xlsx_exc:
                # File may be a legacy binary .xls saved with .xlsx extension — try xlrd
                if "not a zip" in str(xlsx_exc).lower() or "invalid file" in str(xlsx_exc).lower():
                    try:
                        sheets = _xls_to_sheets(source)
                    except Exception as xls_exc:
                        raise Exception(
                            f"Not a valid xlsx or xls file. openpyxl: {xlsx_exc}; xlrd: {xls_exc}"
                        ) from xls_exc
                else:
                    raise
    except Exception as exc:
        out_path = make_output_path(source, output_root, project_root, ".md")
        out_path.write_text(
            f"# Conversion Error: {source.name}\n\n**Error:** `{exc}`\n",
            encoding="utf-8",
        )
        return ConversionResult(
            success=False,
            output_files=[out_path],
            action="error",
            error=str(exc),
        )

    output_files: list[Path] = []
    sheet_summaries: list[str] = []
    out_dir = make_output_dir(source, output_root, project_root)

    for sheet_name, rows in sheets:
        # Sanitize sheet name for filenames
        safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in sheet_name).strip()
        csv_path = out_dir / f"{source.name}.{safe_name}.csv"

        try:
            with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerows(rows)
            output_files.append(csv_path)
        except Exception as exc:
            sheet_summaries.append(f"### Sheet: {sheet_name}\n\n_Error writing CSV: {exc}_\n")
            continue

        preview = _make_md_preview(rows)
        sheet_summaries.append(
            f"### Sheet: {sheet_name}\n\n"
            f"**Rows:** {len(rows) - 1} data rows + 1 header  "
            f"**CSV:** `{csv_path.name}`\n\n"
            f"{preview}\n"
        )

    # --- Markdown summary ---------------------------------------------------
    md_path = make_output_path(source, output_root, project_root, ".md")
    summary_body = "\n\n---\n\n".join(sheet_summaries) if sheet_summaries else "_No sheets found._"
    md_content = (
        f"# Spreadsheet: {source.name}\n\n"
        f"**Sheets:** {len(sheets)}\n\n"
        f"---\n\n"
        f"{summary_body}\n"
    )
    md_path.write_text(md_content, encoding="utf-8")
    output_files.insert(0, md_path)

    return ConversionResult(
        success=True,
        output_files=output_files,
        action="converted",
    )
